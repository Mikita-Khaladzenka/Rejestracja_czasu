from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


class ExcelGenerator:

    HEADER_COLOR = "0078D7"
    ERROR_COLOR = "FF0000"
    HEADER_FONT_COLOR = "FFFFFF"


    @classmethod
    def generate(
        cls,
        report_table: dict,
    ) -> Workbook:


        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Raport"


        styles = cls._create_styles()


        cls._write_table(
            worksheet,
            report_table.get("table", [])
        )


        cls._merge_cells(
            worksheet,
            report_table.get("merged", [])
        )


        cls._style_header(
            worksheet,
            styles
        )


        cls._highlight_errors(
            worksheet,
            report_table.get("red_cells", []),
            styles
        )


        cls._style_cells(
            worksheet,
            styles
        )


        cls._adjust_column_widths(
            worksheet
        )


        return workbook



    @classmethod
    def _create_styles(cls):

        thin = Side(
            style="thin"
        )

        return {

            "header_fill":
                PatternFill(
                    start_color=cls.HEADER_COLOR,
                    end_color=cls.HEADER_COLOR,
                    fill_type="solid"
                ),

            "error_fill":
                PatternFill(
                    start_color=cls.ERROR_COLOR,
                    end_color=cls.ERROR_COLOR,
                    fill_type="solid"
                ),

            "header_font":
                Font(
                    color=cls.HEADER_FONT_COLOR,
                    bold=True
                ),

            "alignment":
                Alignment(
                    horizontal="center",
                    vertical="center"
                ),

            "border":
                Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )
        }



    @staticmethod
    def _write_table(
        worksheet,
        table
    ):

        for row_index, row in enumerate(
            table,
            start=1
        ):

            for column_index, value in enumerate(
                row,
                start=1
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value
                )



    @staticmethod
    def _merge_cells(
        worksheet,
        merged
    ):

        for start_col, start_row, end_col, end_row in merged:

            worksheet.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col
            )



    @staticmethod
    def _style_header(
        worksheet,
        styles
    ):

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=2
        ):

            for cell in row:

                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
                cell.alignment = styles["alignment"]
                cell.border = styles["border"]



    @staticmethod
    def _highlight_errors(
        worksheet,
        red_cells,
        styles
    ):

        for row, column in red_cells:

            worksheet.cell(
                row=row,
                column=column
            ).fill = styles["error_fill"]



    @staticmethod
    def _style_cells(
        worksheet,
        styles
    ):

        for row in worksheet.iter_rows():

            for cell in row:

                cell.alignment = styles["alignment"]
                cell.border = styles["border"]



    @staticmethod
    def _adjust_column_widths(
        worksheet
    ):

        for column in worksheet.columns:

            letter = get_column_letter(
                column[0].column
            )


            width = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column
            )


            worksheet.column_dimensions[
                letter
            ].width = width + 3